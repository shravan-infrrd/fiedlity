import logging
import subprocess as sp
from os import listdir
from os.path import isfile, join
from openpyxl import Workbook
import openpyxl
from os import path
import os
from constant import REFERENCE_FILE
from exceptions.exceptions_handler import InternalServerErrorException

MIN_CHARS_FOR_MACHINE_GENERATED = 50

def update_excel_sheet( result, name):
    if os.path.exists(REFERENCE_FILE):
        wb = openpyxl.load_workbook( REFERENCE_FILE )
    else:
        wb = Workbook()
    try:
        sheet = wb[name]
    except:
        sheet = wb.create_sheet(name)
    for index, (key, value) in enumerate(result.items()):
        if index < 4:
            continue
        sheet.cell(row=1+index, column=1).value = str(key)
        sheet.cell(row=1+index, column=2).value = str(value)

    wb.save(REFERENCE_FILE)
    wb.close

def list_all_files(path):
    onlyfiles = [f for f in listdir(path) if isfile(join(path, f))]
    print(onlyfiles)
    return onlyfiles


def is_machine_generated(pdf_path: str) -> bool:
    #return True
    return False
    #return len(_get_text_from_pdf(pdf_path)) >= MIN_CHARS_FOR_MACHINE_GENERATED


def _get_text_from_pdf(pdf_path: str) -> str:
    extract_text_cmd = [
        "pdftotext",
        pdf_path,
        "-"
    ]
    return sp.check_output(extract_text_cmd)


def formulate_response(result, status, message):
    """
    Formulates the response of all APIs to one response.
    :param result: list of response.
    :return: formulated response with the wrapper.
    """
    formulated_response = {
        "status": {
            "code": "",
            "message": ""
        },
        "data": {

        },
        "errors": []
    }

    try:
        logging.debug("Result = {}".format(result))

        if result is not None:
            formulated_response["data"] = result

        if status is not None:
            formulated_response["status"]["code"] = status
            formulated_response["status"]["message"] = message
        else:
            formulated_response["status"]["code"] = 500
            formulated_response["status"]["message"] = "Something went wrong."

    except Exception as e:
        logging.error("Internal Server Error while forming response ", exc_info=True)
        raise InternalServerErrorException(error_code=500,
                                           error_message=e.__repr__(),
                                           status_code=500, status_message="Internal Server error occured.")

    return formulated_response
