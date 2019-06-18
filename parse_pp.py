import argparse

from utils import list_all_files, update_excel_sheet
#from scanned_to_machined import read_scanned_pdf, read_scanned_image
from controllers.scanned_to_machined import read_scanned_pdf, read_scanned_image
from lib.parse_data import parse_all_fields
from os import path
import os
import uuid

from flask import Flask, request
import json
from flask_pymongo import PyMongo
import copy
import openpyxl
from openpyxl import Workbook

from constant import REFERENCE_FILE, TEST_FILES, TEST_PDF_PATH

app = Flask('mongo')
app.config['MONGO_DBNAME'] = 'cpa_database'
app.config['MONGO_URI'] = 'mongodb://localhost:27017/cpa_database'

mongo = PyMongo(app)


#path = "/Users/shravanc/flask/flask_apps/cpa/UserSamples"

#path = "/Users/shravanc/Desktop/cpa_report/good"
path = "/Users/shravanc/Desktop/CPA_files/new_certificates/data"

files = list_all_files(path)
#files = list_all_files(TEST_FILES)
upload_path = TEST_PDF_PATH #"/Users/shravanc/flask/flask_apps/cpa_certificate_extraction/development/uploads"
print("ALL_FILES====>", files)

wb = Workbook()
"""
def update_excel_sheet(result, name):
		sheet = wb.create_sheet(name)

		for index, (key, value) in enumerate(result.items()):
				if index < 4:
						continue
				sheet.cell(row=1+index, column=1).value = str(key)
				sheet.cell(row=1+index, column=2).value = str(value)

		wb.save(REFERENCE_FILE)
		wb.close
"""
def test_excel_sheet(result, name):
		wb = openpyxl.load_workbook( REFERENCE_FILE )
		try:
				sheet = wb[name]
			 
				for index, (key, value) in enumerate(result.items()):
						if index < 4:
								continue
						key = sheet.cell(row=1+index, column=1).value
						value = sheet.cell(row=1+index, column=2).value
			 
						if value != str(result[key]):
								print(type(value))
								if value == None and str(result[key]) == '':
										continue
								print(f"===ERROR==key->{key}====Expected->{result[key]}===Result->{value}==========FILENAME=====>{name}")
								exit()
				print(f"*{name}*********************************************FINE_SHINE******************************************************")
		except Exception as e:
				print(f"======ERROR====={e}")
				pass

def save_in_db(data):
		certificate_data = copy.deepcopy(data)
		mongo.db.certificates.insert(certificate_data)

"""
parser = argparse.ArgumentParser()
parser.add_argument("o",help="operation to execute (update / test) ")
parser.add_argument("f", help="File Name to execute the operation given above. If all files then pass 'all'")
args = parser.parse_args()
print(args.echo)

operation = args.o
input_file_name = args.f

if input_file_name == 'all':
"""

file = open("program_name.txt", 'w')

for fp in files:
		print("fp---->", fp, os.path.basename(fp).split('.')[0])
		extension = os.path.basename(fp).split('.')[1]
		if extension in ['png', 'jpg', 'jpeg']:
				continue
		if fp.startswith('.'):
				continue
		file_name_without_ext = os.path.basename(fp).split('.')[0]
			
		filename = path + '/' + fp
		doc_dir_location = os.path.join(upload_path, file_name_without_ext)
		if not os.path.exists(doc_dir_location):
				os.makedirs(doc_dir_location)
		result = read_scanned_pdf( filename, doc_dir_location)
	 
	 
		text_file_path = os.path.join(upload_path, file_name_without_ext, 'texts', 'stitched.txt')
	
		try: 
				with open( text_file_path ) as fp1:
						contents = fp1.readlines()
				parse_all_fields(contents, result)
				string = result["program_name"] + "****" + fp + "\n\n"
				file.write(string)
				print("Result**************", result)
		except:
				pass
		#save_in_db(result) 
		#update_excel_sheet(result, file_name_without_ext)
		#test_excel_sheet(result, fp)
		print("===COMPLETED===")

file.close()
