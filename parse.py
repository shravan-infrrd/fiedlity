import argparse
from time import sleep

from utils import list_all_files, update_excel_sheet
#from scanned_to_machined import read_scanned_pdf, read_scanned_image
from controllers.scanned_to_machined import read_scanned_pdf, read_scanned_image
from lib.parse_certificate import parse_all_fields
from os import path
import os
import uuid

from flask import Flask, request
import json
from flask_pymongo import PyMongo
import copy
import openpyxl
from openpyxl import Workbook

from constant import REFERENCE_FILE, TEST_FILES, TEST_PDF_PATH

app = Flask('mongo')
app.config['MONGO_DBNAME'] = 'cpa_database'
app.config['MONGO_URI'] = 'mongodb://localhost:27017/cpa_database'

mongo = PyMongo(app)


#path = "/Users/shravanc/flask/flask_apps/cpa/UserSamples"

path = "/Users/shravanc/Desktop/cpa_report/good"
path = "/Users/shravanc/Desktop/CPA_files/new_certificates/data"
files = list_all_files(path)
#files = list_all_files(TEST_FILES)
upload_path = TEST_PDF_PATH #"/Users/shravanc/flask/flask_apps/cpa_certificate_extraction/development/uploads"
print("ALL_FILES====>", files)

wb = Workbook()
"""
def update_excel_sheet(result, name):
		sheet = wb.create_sheet(name)

		for index, (key, value) in enumerate(result.items()):
				if index < 4:
						continue
				sheet.cell(row=1+index, column=1).value = str(key)
				sheet.cell(row=1+index, column=2).value = str(value)

		wb.save(REFERENCE_FILE)
		wb.close
"""
def test_excel_sheet(result, name):
		wb = openpyxl.load_workbook( REFERENCE_FILE )
		try:
				sheet = wb[name]
			 
				for index, (key, value) in enumerate(result.items()):
						if index < 4:
								continue
						key = sheet.cell(row=1+index, column=1).value
						value = sheet.cell(row=1+index, column=2).value
			 
						if value != str(result[key]):
								print(type(value))
								if value == None and str(result[key]) == '':
										continue
								print(f"===ERROR==key->{key}====Expected->{result[key]}===Result->{value}==========FILENAME=====>{name}")
								exit()
				print(f"*{name}*********************************************FINE_SHINE******************************************************")
		except Exception as e:
				print(f"======ERROR====={e}")
				pass

def save_in_db(data):
		certificate_data = copy.deepcopy(data)
		mongo.db.certificates.insert(certificate_data)

"""
parser = argparse.ArgumentParser()
parser.add_argument("o",help="operation to execute (update / test) ")
parser.add_argument("f", help="File Name to execute the operation given above. If all files then pass 'all'")
args = parser.parse_args()
print(args.echo)

operation = args.o
input_file_name = args.f

if input_file_name == 'all':
"""

name_file = open("name.txt", 'w')
pn_file = open("program_name.txt", 'w')
dm_file = open("delivery_method.txt", 'w')
date_file = open("date.txt", 'w')
sponsor_file = open("sponsor.txt", 'w')
sponsor_id_file = open("sponsor_id.txt", 'w')
fos_file = open("fos.txt", "w")

name_count = 0
program_name_count = 0
delivery_method_count = 0
date_count = 0


min_range = 230
max_range = 100

for index, fp in enumerate(files):
		flag = True #False
		#for fname in invalid_list:
		#		if fp == fname:
		#				flag = True

		print(f"0_index--->{index}")
		if index <=min_range:
				print(f"1_index--->{index}")
				continue
		elif index > max_range:
				print(f"2_index--->{index}")
				exit()
		else:
				print(f"3_index--->{index}")
				pass

	
		if flag:
				#if index < 2000  and index > 1000:
				#		pass
				#elif index > 2000:
				#		exit()
				#else:
				#		continue
				print("fp---->", fp, os.path.basename(fp).split('.')[0])
				extension = os.path.basename(fp).split('.')[1]
				if extension in ['png', 'jpg', 'jpeg']:
						continue
				if fp.startswith('.'):
						continue
				file_name_without_ext = os.path.basename(fp).split('.')[0]
					
				filename = path + '/' + fp
				doc_dir_location = os.path.join(upload_path, file_name_without_ext)
				if not os.path.exists(doc_dir_location):
						os.makedirs(doc_dir_location)
				result = read_scanned_pdf( filename, doc_dir_location)
	     
				try:	 
						text_file_path = os.path.join(upload_path, file_name_without_ext, 'texts', 'stitched.txt')
						
						with open( text_file_path ) as fp1:
								contents = fp1.readlines()
            
						result = parse_all_fields(contents, {})
           
						print("RESULT**************************>", result) 
						if result["username"] == "":
								name_count = name_count + 1
						if result["program_name"] == "":
								program_name_count = program_name_count + 1
						if result['delivery_format'] == "":
								delivery_method_count = delivery_method_count + 1
						if result['completion_date'] == "":
								date_count = date_count + 1
            
            
            
						print("result--->", result)
						name_string = result["username"]['value'] #+ "****" + fp #+ "\n"	
						pn_string =   result['program_name']['value'] #+ "****" + fp #+ "\n\n"
						dm_string =   result['delivery_format']['value'] #+ "****" + fp #+ "\n\n"
						date_string =   result['completion_date']['value'] #+ "****" + fp #+ "\n\n"
						try:
								sponsor_string = result['sponsor_name'][0]['value'] #+ "****" + fp #+ "\n\n"
						except:
								sponsor_string = ""
						try:
								sponsor_id_string = result['sponsor_number'][0]['value'] #+ "****" + fp #+ "\n\n"
						except:
								sponsor_id_string = ""
						try:
								fos_string = result["field_of_study"][0]["name"]
						except:
								fos_string = ""
        
        
						print(f"1*****************FOS_STRING***********************|{fos_string.strip()}|******************************{len(fos_string.strip())}, =====>")
						print(f"2*****************DATE_STRING***********************|{date_string.strip()}|******************************{len(date_string.strip())}, =====>")
						#print(f"1*****************PROGRAMNAME***********************|{pn_string.strip()}|******************************{len(pn_string.strip())}")
						#pn_file.write("--name--" + pn_string.strip() + "--\n")
						if name_string.strip() == "":
								name_file.write(name_string + "****" + fp + "\n")
						#print(f"2*****************PROGRAMNAME***********************|{pn_string}|******************************pn_string.strip() == """)
						if pn_string.strip() == "":
								#print(f"3*****************PROGRAMNAME***********************|{pn_string}|******************************")
								pn_file.write(pn_string + "****" + fp + "\n")
						if dm_string.strip() == "":
								dm_file.write(dm_string + "****" + fp + "\n")
						if date_string.strip() == "":
								date_file.write(date_string + "****" + fp + "\n")
						if sponsor_string.strip() == "":
								sponsor_file.write(sponsor_string + "****" + fp + "\n")
						if sponsor_id_string.strip() == "":
								sponsor_id_file.write(sponsor_id_string + "****" + fp + "\n")
						if fos_string.strip() == "":
								print(f"*****************FOS_STRING***********************|{fos_string.strip()}|******************************{len(fos_string.strip())}")
								fos_file.write( fos_string + "****" + fp + "\n" )

						#save_in_db(result) 
						#update_excel_sheet(result, file_name_without_ext)
						#test_excel_sheet(result, fp)
						sleep(0.03)
						print("===COMPLETED===")	
	    
				except Exception as error:
						print(f"Error----->{error}")
						continue

name_file.close()
pn_file.close()
dm_file.close()
print(f"name_count = {name_count}, date_count = {date_count}, program_name_count = {program_name_count}, delivery_method_count = {delivery_method_count}")
